import os
import random
from io import BytesIO
from datetime import datetime, timedelta

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from flask import Flask, request, jsonify, render_template, send_file, send_from_directory

app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(__file__), '..', 'templates'),
    static_folder=os.path.join(os.path.dirname(__file__), '..', 'static'),
)


# ─────────────────────────────────────────────────────────────────────────────
# Latin Square generation
# ─────────────────────────────────────────────────────────────────────────────

def generate_latin_square(n: int, seed: int = 42) -> list:
    """
    Generate an n×n Latin Square using cyclic construction with random
    row and symbol permutations. Each symbol (team index) appears exactly
    once per row and once per column. Guaranteed by construction.

    Tries multiple random permutations and picks the one where teams
    move around the most between consecutive rows (least "same column" repeats).
    """
    best = None
    best_score = float("inf")

    for attempt in range(200):
        rng = random.Random(seed * 1000 + attempt)

        sym = list(range(n))
        rng.shuffle(sym)
        rows = list(range(n))
        rng.shuffle(rows)

        square = [[sym[(rows[i] + j) % n] for j in range(n)] for i in range(n)]

        # Penalise team staying in same column position between consecutive rows
        score = 0
        for i in range(1, n):
            for j in range(n):
                team = square[i][j]
                prev_col = square[i - 1].index(team)
                if prev_col == j:
                    score += 1

        if score < best_score:
            best_score = score
            best = square
            if score == 0:
                break

    return best


def _all_pairings(items):
    """Yield every way to partition items into pairs (len must be even)."""
    if len(items) == 0:
        yield []
        return
    first = items[0]
    for i, second in enumerate(items[1:], 1):
        pair = (first, second)
        remaining = items[1:i] + items[i + 1:]
        for sub in _all_pairings(remaining):
            yield [pair] + sub


def optimize_column_pairing(square, n_duo, n_solo, seed=42):
    """
    Reorder the columns of the Latin Square so that duo-game column pairs
    produce maximum opponent diversity. Tries different column pairings
    and picks the one where each team faces the most unique opponents.

    The Latin Square properties are preserved under column permutation.
    """
    n = len(square)
    if n_duo <= 1:
        return square

    duo_cols = list(range(2 * n_duo))  # columns used for duo games

    def score_pairing(pairing):
        """Higher = better. Returns (min_unique, total_unique)."""
        opponents = {t: set() for t in range(n)}
        for i in range(n):
            for c1, c2 in pairing:
                t1, t2 = square[i][c1], square[i][c2]
                opponents[t1].add(t2)
                opponents[t2].add(t1)
        counts = [len(v) for v in opponents.values()]
        return (min(counts), sum(counts))

    # For small column counts, try all possible pairings exhaustively
    if 2 * n_duo <= 10:
        best_pairing = None
        best_score = (-1, -1)
        for pairing in _all_pairings(duo_cols):
            s = score_pairing(pairing)
            if s > best_score:
                best_score = s
                best_pairing = pairing
    else:
        # For larger n, sample random pairings
        best_pairing = list(zip(duo_cols[::2], duo_cols[1::2]))
        best_score = score_pairing(best_pairing)
        for attempt in range(2000):
            rng = random.Random(seed * 3000 + attempt)
            shuffled = duo_cols[:]
            rng.shuffle(shuffled)
            pairing = list(zip(shuffled[::2], shuffled[1::2]))
            s = score_pairing(pairing)
            if s > best_score:
                best_score = s
                best_pairing = pairing

    # Build the column reorder: duo pairs first (in pairing order), then solo cols
    col_order = []
    for c1, c2 in best_pairing:
        col_order.extend([c1, c2])
    col_order.extend(range(2 * n_duo, n))  # solo columns stay at the end

    return [[row[c] for c in col_order] for row in square]


# ─────────────────────────────────────────────────────────────────────────────
# Build display DataFrame from Latin Square + game type info
# ─────────────────────────────────────────────────────────────────────────────

def build_dataframe(square, team_names, duo_names, solo_names, slot_labels, n_duo, n_solo):
    """
    The Latin Square has n columns (positions). We group them:
      - Positions 0..2*n_duo-1 → n_duo duo game columns (pairs of 2)
      - Positions 2*n_duo..n-1 → n_solo solo game columns (1 each)

    Duo columns show "Team A vs Team B", solo columns show "Team A".
    """
    col_headers = list(duo_names) + list(solo_names)
    rows = []
    for i in range(len(slot_labels)):
        row = []
        # Duo games: pair adjacent positions
        for d in range(n_duo):
            t1 = team_names[square[i][2 * d]]
            t2 = team_names[square[i][2 * d + 1]]
            row.append(f"{t1} vs {t2}")
        # Solo games
        for s in range(n_solo):
            t = team_names[square[i][2 * n_duo + s]]
            row.append(t)
        rows.append(row)
    df = pd.DataFrame(rows, index=slot_labels, columns=col_headers)
    df.index.name = "Tijdslot"
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Time labels
# ─────────────────────────────────────────────────────────────────────────────

def make_time_labels(n_slots, start, game_min, switch_min, break_after=None, break_min=0):
    labels = []
    t = start
    for i in range(n_slots):
        end = t + timedelta(minutes=game_min)
        labels.append(f"{t.strftime('%H:%M')} – {end.strftime('%H:%M')}")
        t = end + timedelta(minutes=switch_min)
        if break_after and break_min > 0 and (i + 1) == break_after:
            t += timedelta(minutes=break_min)
    return labels


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill):
    """Auto-size columns and set row heights for a worksheet."""
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(w + 4, 14)
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24


def _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill,
                        break_after=None, break_min=0):
    """Write a simple table with header row + data rows (first col = index style).
    Optionally inserts a styled break row after break_after data rows."""
    break_fill = PatternFill("solid", fgColor="FEF3C7")
    break_font = Font(bold=True, color="92400E", size=11)
    n_cols = len(headers)

    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr

    ws_row = 2
    for i, row in enumerate(rows):
        row_fill = even_fill if ws_row % 2 == 0 else odd_fill
        for j, val in enumerate(row, 1):
            c = ws.cell(ws_row, j, val)
            if j == 1:
                c.font = white_f; c.fill = idx_fill
            else:
                c.font = body_f; c.fill = row_fill
            c.alignment = center; c.border = bdr
        ws_row += 1

        # Insert break row after break_after rounds (1-indexed)
        if break_after and break_min > 0 and (i + 1) == break_after:
            if n_cols > 1:
                ws.merge_cells(start_row=ws_row, start_column=1, end_row=ws_row, end_column=n_cols)
            c = ws.cell(ws_row, 1, f"☕  PAUZE — {break_min} minuten")
            c.font = break_font; c.fill = break_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[ws_row].height = 24
            ws_row += 1

    _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)


def to_excel(df, n_duo, square, slot_labels, team_names, duo_names, solo_names, n_solo,
             game_descriptions=None, break_after=None, break_min=0):
    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    idx_fill  = PatternFill("solid", fgColor="2E75B6")
    solo_fill = PatternFill("solid", fgColor="FFE699")
    even_fill = PatternFill("solid", fgColor="DEEAF1")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    white_f   = Font(bold=True, color="FFFFFF", size=11)
    dark_f    = Font(bold=True, color="1F4E79", size=11)
    body_f    = Font(size=11)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="BFBFBF")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    n = len(square)

    # ── Tab 1: Overall schema ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Schema"

    c = ws.cell(1, 1, df.index.name or "Tijdslot")
    c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr
    for j, col in enumerate(df.columns, 2):
        is_solo = (j - 2) >= n_duo
        c = ws.cell(1, j, col)
        c.font = dark_f if is_solo else white_f
        c.fill = solo_fill if is_solo else hdr_fill
        c.alignment = center; c.border = bdr

    break_fill = PatternFill("solid", fgColor="FEF3C7")
    break_font = Font(bold=True, color="92400E", size=11)
    n_cols = len(df.columns) + 1  # +1 for index column

    ws_row = 2
    for i, (idx_val, row) in enumerate(df.iterrows()):
        row_fill = even_fill if ws_row % 2 == 0 else odd_fill
        ic = ws.cell(ws_row, 1, idx_val)
        ic.font = white_f; ic.fill = idx_fill; ic.alignment = center; ic.border = bdr
        for j, val in enumerate(row, 2):
            is_solo = (j - 2) >= n_duo
            c = ws.cell(ws_row, j, val)
            c.font = body_f
            c.fill = solo_fill if is_solo else row_fill
            c.alignment = center; c.border = bdr
        ws_row += 1

        if break_after and break_min > 0 and (i + 1) == break_after:
            ws.merge_cells(start_row=ws_row, start_column=1, end_row=ws_row, end_column=n_cols)
            c = ws.cell(ws_row, 1, f"☕  PAUZE — {break_min} minuten")
            c.font = break_font; c.fill = break_fill
            c.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[ws_row].height = 24
            ws_row += 1

    _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)

    # ── Tabs per game ────────────────────────────────────────────────────────
    all_game_names = list(duo_names) + list(solo_names)
    for g_idx, game_name in enumerate(all_game_names):
        ws = wb.create_sheet(title=game_name[:31])  # Excel tab name max 31 chars
        is_duo = g_idx < n_duo

        if is_duo:
            headers = ["Tijdslot", "Team 1", "Team 2", "Score"]
            rows = []
            for i, label in enumerate(slot_labels):
                t1 = team_names[square[i][2 * g_idx]]
                t2 = team_names[square[i][2 * g_idx + 1]]
                rows.append([label, t1, t2, ""])
        else:
            solo_idx = g_idx - n_duo
            headers = ["Tijdslot", "Team", "Score"]
            rows = []
            for i, label in enumerate(slot_labels):
                t = team_names[square[i][2 * n_duo + solo_idx]]
                rows.append([label, t, ""])

        _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill,
                            break_after=break_after, break_min=break_min)

        # Add description below the table
        if game_descriptions and game_descriptions.get(game_name, "").strip():
            desc_row = len(rows) + 3  # 1 header + len(rows) data + 1 blank row
            c = ws.cell(desc_row, 1, "Beschrijving")
            c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr
            desc_row += 1
            desc_text = game_descriptions[game_name].strip()
            c = ws.cell(desc_row, 1, desc_text)
            c.font = body_f
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Merge across all columns for readability
            n_hdr_cols = len(headers)
            if n_hdr_cols > 1:
                ws.merge_cells(start_row=desc_row, start_column=1, end_row=desc_row, end_column=n_hdr_cols)
            # Auto-height for description
            line_count = desc_text.count("\n") + 1
            ws.row_dimensions[desc_row].height = max(24, 16 * line_count)

    # ── Tabs per team ────────────────────────────────────────────────────────
    for tm_idx, tm_name in enumerate(team_names):
        ws = wb.create_sheet(title=tm_name[:31])
        headers = ["Tijdslot", "Spel", "Tegenstander"]
        rows = []
        for i, label in enumerate(slot_labels):
            pos = square[i].index(tm_idx)
            if pos < 2 * n_duo:
                game_idx = pos // 2
                partner_pos = pos + 1 if pos % 2 == 0 else pos - 1
                partner = team_names[square[i][partner_pos]]
                rows.append([label, duo_names[game_idx], partner])
            else:
                solo_idx = pos - 2 * n_duo
                rows.append([label, solo_names[solo_idx], ""])

        _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill,
                            break_after=break_after, break_min=break_min)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Flask routes
# ─────────────────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    return render_template("index.html")


@app.route("/favicon.ico")
def favicon():
    return send_from_directory(
        os.path.join(os.path.dirname(__file__), '..', 'static'),
        'favicon.svg',
        mimetype='image/svg+xml'
    )


@app.route("/api/generate", methods=["POST"])
def api_generate():
    data = request.get_json(force=True)

    n           = int(data.get("n", 6))
    seed        = int(data.get("seed", 42))
    n_duo       = int(data.get("n_duo", 3))
    n_solo      = int(data.get("n_solo", 0))
    team_names  = data.get("team_names", [f"Team {i+1}" for i in range(n)])
    duo_names   = data.get("duo_names", [f"Duo spel {i+1}" for i in range(n_duo)])
    solo_names  = data.get("solo_names", [f"Solo spel {i+1}" for i in range(n_solo)])
    start_time  = data.get("start_time", "09:00")
    game_min    = int(data.get("game_min", 20))
    switch_min  = int(data.get("switch_min", 5))
    break_after = int(data["break_after"]) if data.get("break_after") else None
    break_min   = int(data.get("break_min", 0))

    # Validate
    errors = []
    if n < 2 or n > 20:
        errors.append("Aantal teams moet tussen 2 en 20 liggen.")
    if n_duo < 0:
        errors.append("Aantal duo-spellen mag niet negatief zijn.")
    if n_solo < 0:
        errors.append("Aantal solo-spellen mag niet negatief zijn.")
    if len(team_names) != n:
        errors.append("Aantal teamnamen klopt niet met het aantal teams.")
    if len(set(team_names)) < n:
        errors.append("Teamnamen moeten uniek zijn.")
    if 2 * n_duo + n_solo != n:
        errors.append(f"duo×2 + solo moet gelijk zijn aan n teams: {2*n_duo} + {n_solo} ≠ {n}")
    all_game_names = duo_names + solo_names
    if len(set(all_game_names)) < len(all_game_names):
        errors.append("Spelnamen moeten uniek zijn.")
    if errors:
        return jsonify({"error": "; ".join(errors)}), 400

    # Parse start time
    try:
        start_dt = datetime.strptime(start_time, "%H:%M")
    except ValueError:
        return jsonify({"error": "Ongeldig tijdformaat. Gebruik HH:MM."}), 400

    # Generate
    square = generate_latin_square(n, seed=seed)
    square = optimize_column_pairing(square, n_duo, n_solo, seed=seed)
    slot_labels = make_time_labels(n, start_dt, game_min, switch_min, break_after=break_after, break_min=break_min)
    df = build_dataframe(square, team_names, duo_names, solo_names, slot_labels, n_duo, n_solo)

    # Compute end time (including break)
    total_min = n * (game_min + switch_min) - switch_min + (break_min if break_after and break_min else 0)
    end_dt = start_dt + timedelta(minutes=total_min)

    # Latin square display values (1-based)
    ls_rows = [[square[i][j] + 1 for j in range(n)] for i in range(n)]
    ls_col_sums = [sum(square[i][j] + 1 for i in range(n)) for j in range(n)]
    ls_row_sums = [sum(square[i][j] + 1 for j in range(n)) for i in range(n)]
    expected_sum = n * (n + 1) // 2

    # Per-team schedule (server-side)
    team_schedules = {}
    for tm_idx, tm_name in enumerate(team_names):
        schedule = []
        for i, label in enumerate(slot_labels):
            pos = square[i].index(tm_idx)
            if pos < 2 * n_duo:
                game_idx = pos // 2
                partner_pos = pos + 1 if pos % 2 == 0 else pos - 1
                partner = team_names[square[i][partner_pos]]
                schedule.append({
                    "slot": label,
                    "game": duo_names[game_idx],
                    "opponent": partner,
                })
            else:
                solo_idx_val = pos - 2 * n_duo
                schedule.append({
                    "slot": label,
                    "game": solo_names[solo_idx_val],
                    "opponent": "",
                })
        team_schedules[tm_name] = schedule

    return jsonify({
        "square":       square,
        "slot_labels":  slot_labels,
        "df_columns":   list(df.columns),
        "df_rows":      df.values.tolist(),
        "n":            n,
        "n_duo":        n_duo,
        "n_solo":       n_solo,
        "team_names":   team_names,
        "duo_names":    duo_names,
        "solo_names":   solo_names,
        "end_time":     end_dt.strftime("%H:%M"),
        "total_min":    total_min,
        "expected_sum": expected_sum,
        "ls_rows":      ls_rows,
        "ls_col_sums":  ls_col_sums,
        "ls_row_sums":  ls_row_sums,
        "team_schedules": team_schedules,
        "break_after":  break_after,
        "break_min":    break_min,
    })


@app.route("/api/download", methods=["POST"])
def api_download():
    data = request.get_json(force=True)

    square           = data.get("square")
    slot_labels      = data.get("slot_labels")
    df_columns       = data.get("df_columns")
    df_rows          = data.get("df_rows")
    n_duo            = int(data.get("n_duo", 0))
    n_solo           = int(data.get("n_solo", 0))
    team_names       = data.get("team_names", [])
    duo_names        = data.get("duo_names", [])
    solo_names       = data.get("solo_names", [])
    game_descriptions = data.get("game_descriptions", {})
    file_name        = data.get("file_name", "spellenschema")
    break_after      = int(data["break_after"]) if data.get("break_after") else None
    break_min        = int(data.get("break_min", 0))

    if not square or not slot_labels or not df_columns or df_rows is None:
        return jsonify({"error": "Ontbrekende gegevens voor download."}), 400

    # Reconstruct DataFrame
    df = pd.DataFrame(df_rows, index=slot_labels, columns=df_columns)
    df.index.name = "Tijdslot"

    excel_bytes = to_excel(
        df, n_duo, square, slot_labels,
        team_names, duo_names, solo_names, n_solo,
        game_descriptions, break_after=break_after, break_min=break_min,
    )

    safe_name = file_name.strip() or "spellenschema"
    if not safe_name.endswith(".xlsx"):
        safe_name += ".xlsx"

    return send_file(
        BytesIO(excel_bytes),
        as_attachment=True,
        download_name=safe_name,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 5000))
    app.run(debug=True, port=port)
