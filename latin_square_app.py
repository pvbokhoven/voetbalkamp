import streamlit as st
import pandas as pd
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta
import random


# ─────────────────────────────────────────────────────────────────────────────
# Latin Square generation
# ─────────────────────────────────────────────────────────────────────────────

def generate_latin_square(n: int, seed: int = 42) -> list[list[int]]:
    """
    Generate an n×n Latin Square using cyclic construction with random
    row and symbol permutations. Each symbol (team index) appears exactly
    once per row and once per column. Guaranteed by construction.

    Tries multiple random permutations and picks the one where teams
    move around the most between consecutive rows (least "same column" repeats).
    """
    best = None
    best_score = float("inf")

    for attempt in range(200):
        rng = random.Random(seed * 1000 + attempt)

        sym = list(range(n))
        rng.shuffle(sym)
        rows = list(range(n))
        rng.shuffle(rows)

        square = [[sym[(rows[i] + j) % n] for j in range(n)] for i in range(n)]

        # Penalise team staying in same column position between consecutive rows
        score = 0
        for i in range(1, n):
            for j in range(n):
                team = square[i][j]
                prev_col = square[i - 1].index(team)
                if prev_col == j:
                    score += 1

        if score < best_score:
            best_score = score
            best = square
            if score == 0:
                break

    return best


def _all_pairings(items):
    """Yield every way to partition items into pairs (len must be even)."""
    if len(items) == 0:
        yield []
        return
    first = items[0]
    for i, second in enumerate(items[1:], 1):
        pair = (first, second)
        remaining = items[1:i] + items[i + 1:]
        for sub in _all_pairings(remaining):
            yield [pair] + sub


def optimize_column_pairing(square, n_duo, n_solo, seed=42):
    """
    Reorder the columns of the Latin Square so that duo-game column pairs
    produce maximum opponent diversity. Tries different column pairings
    and picks the one where each team faces the most unique opponents.

    The Latin Square properties are preserved under column permutation.
    """
    n = len(square)
    if n_duo <= 1:
        return square

    duo_cols = list(range(2 * n_duo))  # columns used for duo games

    def score_pairing(pairing):
        """Higher = better. Returns (min_unique, total_unique)."""
        opponents = {t: set() for t in range(n)}
        for i in range(n):
            for c1, c2 in pairing:
                t1, t2 = square[i][c1], square[i][c2]
                opponents[t1].add(t2)
                opponents[t2].add(t1)
        counts = [len(v) for v in opponents.values()]
        return (min(counts), sum(counts))

    # For small column counts, try all possible pairings exhaustively
    if 2 * n_duo <= 10:
        best_pairing = None
        best_score = (-1, -1)
        for pairing in _all_pairings(duo_cols):
            s = score_pairing(pairing)
            if s > best_score:
                best_score = s
                best_pairing = pairing
    else:
        # For larger n, sample random pairings
        best_pairing = list(zip(duo_cols[::2], duo_cols[1::2]))
        best_score = score_pairing(best_pairing)
        for attempt in range(2000):
            rng = random.Random(seed * 3000 + attempt)
            shuffled = duo_cols[:]
            rng.shuffle(shuffled)
            pairing = list(zip(shuffled[::2], shuffled[1::2]))
            s = score_pairing(pairing)
            if s > best_score:
                best_score = s
                best_pairing = pairing

    # Build the column reorder: duo pairs first (in pairing order), then solo cols
    col_order = []
    for c1, c2 in best_pairing:
        col_order.extend([c1, c2])
    col_order.extend(range(2 * n_duo, n))  # solo columns stay at the end

    return [[row[c] for c in col_order] for row in square]


# ─────────────────────────────────────────────────────────────────────────────
# Build display DataFrame from Latin Square + game type info
# ─────────────────────────────────────────────────────────────────────────────

def build_dataframe(square, team_names, duo_names, solo_names, slot_labels, n_duo, n_solo):
    """
    The Latin Square has n columns (positions). We group them:
      - Positions 0..2*n_duo-1 → n_duo duo game columns (pairs of 2)
      - Positions 2*n_duo..n-1 → n_solo solo game columns (1 each)

    Duo columns show "Team A vs Team B", solo columns show "Team A".
    """
    col_headers = list(duo_names) + list(solo_names)
    rows = []
    for i in range(len(slot_labels)):
        row = []
        # Duo games: pair adjacent positions
        for d in range(n_duo):
            t1 = team_names[square[i][2 * d]]
            t2 = team_names[square[i][2 * d + 1]]
            row.append(f"{t1} vs {t2}")
        # Solo games
        for s in range(n_solo):
            t = team_names[square[i][2 * n_duo + s]]
            row.append(t)
        rows.append(row)
    df = pd.DataFrame(rows, index=slot_labels, columns=col_headers)
    df.index.name = "Tijdslot"
    return df


# ─────────────────────────────────────────────────────────────────────────────
# Time labels
# ─────────────────────────────────────────────────────────────────────────────

def make_time_labels(n_slots, start, game_min, switch_min):
    labels = []
    t = start
    for _ in range(n_slots):
        end = t + timedelta(minutes=game_min)
        labels.append(f"{t.strftime('%H:%M')} – {end.strftime('%H:%M')}")
        t = end + timedelta(minutes=switch_min)
    return labels


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill):
    """Auto-size columns and set row heights for a worksheet."""
    for col_cells in ws.columns:
        w = max((len(str(c.value or "")) for c in col_cells), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = max(w + 4, 14)
    for r in range(1, ws.max_row + 1):
        ws.row_dimensions[r].height = 24


def _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill):
    """Write a simple table with header row + data rows (first col = index style)."""
    for j, h in enumerate(headers, 1):
        c = ws.cell(1, j, h)
        c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr
    for i, row in enumerate(rows, 2):
        row_fill = even_fill if i % 2 == 0 else odd_fill
        for j, val in enumerate(row, 1):
            c = ws.cell(i, j, val)
            if j == 1:
                c.font = white_f; c.fill = idx_fill
            else:
                c.font = body_f; c.fill = row_fill
            c.alignment = center; c.border = bdr
    _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)


def to_excel(df, n_duo, square, slot_labels, team_names, duo_names, solo_names, n_solo, game_descriptions=None):
    wb = openpyxl.Workbook()

    hdr_fill  = PatternFill("solid", fgColor="1F4E79")
    idx_fill  = PatternFill("solid", fgColor="2E75B6")
    solo_fill = PatternFill("solid", fgColor="FFE699")
    even_fill = PatternFill("solid", fgColor="DEEAF1")
    odd_fill  = PatternFill("solid", fgColor="FFFFFF")
    white_f   = Font(bold=True, color="FFFFFF", size=11)
    dark_f    = Font(bold=True, color="1F4E79", size=11)
    body_f    = Font(size=11)
    center    = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin      = Side(style="thin", color="BFBFBF")
    bdr       = Border(left=thin, right=thin, top=thin, bottom=thin)

    n = len(square)

    # ── Tab 1: Overall schema ────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Schema"

    c = ws.cell(1, 1, df.index.name or "Tijdslot")
    c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr
    for j, col in enumerate(df.columns, 2):
        is_solo = (j - 2) >= n_duo
        c = ws.cell(1, j, col)
        c.font = dark_f if is_solo else white_f
        c.fill = solo_fill if is_solo else hdr_fill
        c.alignment = center; c.border = bdr

    for i, (idx_val, row) in enumerate(df.iterrows(), 2):
        row_fill = even_fill if i % 2 == 0 else odd_fill
        ic = ws.cell(i, 1, idx_val)
        ic.font = white_f; ic.fill = idx_fill; ic.alignment = center; ic.border = bdr
        for j, val in enumerate(row, 2):
            is_solo = (j - 2) >= n_duo
            c = ws.cell(i, j, val)
            c.font = body_f
            c.fill = solo_fill if is_solo else row_fill
            c.alignment = center; c.border = bdr

    _style_sheet(ws, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)

    # ── Tabs per game ────────────────────────────────────────────────────────
    all_game_names = list(duo_names) + list(solo_names)
    for g_idx, game_name in enumerate(all_game_names):
        ws = wb.create_sheet(title=game_name[:31])  # Excel tab name max 31 chars
        is_duo = g_idx < n_duo

        if is_duo:
            headers = ["Tijdslot", "Team 1", "Team 2", "Score"]
            rows = []
            for i, label in enumerate(slot_labels):
                t1 = team_names[square[i][2 * g_idx]]
                t2 = team_names[square[i][2 * g_idx + 1]]
                rows.append([label, t1, t2, ""])
        else:
            solo_idx = g_idx - n_duo
            headers = ["Tijdslot", "Team", "Score"]
            rows = []
            for i, label in enumerate(slot_labels):
                t = team_names[square[i][2 * n_duo + solo_idx]]
                rows.append([label, t, ""])

        _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)

        # Add description below the table
        if game_descriptions and game_descriptions.get(game_name, "").strip():
            desc_row = len(rows) + 3  # 1 header + len(rows) data + 1 blank row
            c = ws.cell(desc_row, 1, "Beschrijving")
            c.font = white_f; c.fill = hdr_fill; c.alignment = center; c.border = bdr
            desc_row += 1
            desc_text = game_descriptions[game_name].strip()
            c = ws.cell(desc_row, 1, desc_text)
            c.font = body_f
            c.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            # Merge across all columns for readability
            n_hdr_cols = len(headers)
            if n_hdr_cols > 1:
                ws.merge_cells(start_row=desc_row, start_column=1, end_row=desc_row, end_column=n_hdr_cols)
            # Auto-height for description
            line_count = desc_text.count("\n") + 1
            ws.row_dimensions[desc_row].height = max(24, 16 * line_count)

    # ── Tabs per team ────────────────────────────────────────────────────────
    for tm_idx, tm_name in enumerate(team_names):
        ws = wb.create_sheet(title=tm_name[:31])
        headers = ["Tijdslot", "Spel", "Tegenstander"]
        rows = []
        for i, label in enumerate(slot_labels):
            pos = square[i].index(tm_idx)
            if pos < 2 * n_duo:
                game_idx = pos // 2
                partner_pos = pos + 1 if pos % 2 == 0 else pos - 1
                partner = team_names[square[i][partner_pos]]
                rows.append([label, duo_names[game_idx], partner])
            else:
                solo_idx = pos - 2 * n_duo
                rows.append([label, solo_names[solo_idx], ""])

        _write_simple_table(ws, headers, rows, hdr_fill, idx_fill, white_f, body_f, center, bdr, even_fill, odd_fill)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


# ─────────────────────────────────────────────────────────────────────────────
# Streamlit
# ─────────────────────────────────────────────────────────────────────────────

def main():
    st.set_page_config(page_title="Voetbalkamp Schema", layout="wide")

    # ── Sidebar ───────────────────────────────────────────────────────────────
    with st.sidebar:
        try:
            st.image("logo-stiphout-vooruit.svg", width=90)
        except Exception:
            pass

        st.title("Instellingen")

        # Teams
        st.subheader("Teams")
        n = int(st.number_input("Aantal teams", min_value=2, max_value=20, value=6, step=1))

        use_def_teams = st.toggle("Standaard teamnamen", value=True, key="tog_teams")
        team_names: list[str] = []
        if use_def_teams:
            team_names = [f"Team {i + 1}" for i in range(n)]
            st.caption(", ".join(team_names))
        else:
            for i in range(n):
                nm = st.text_input(f"Naam team {i + 1}", value=f"Team {i + 1}", key=f"team_{i}")
                team_names.append(nm.strip() or f"Team {i + 1}")

        st.divider()

        # Game types
        st.subheader("Speltypen")
        max_duo = n // 2
        n_duo = int(st.number_input(
            "Aantal duo-spellen (2 teams tegen elkaar)",
            min_value=0, max_value=max_duo, value=max_duo, step=1,
        ))
        n_solo = n - 2 * n_duo
        st.caption(f"→ **{n_duo}** duo-spellen × 2 + **{n_solo}** solo-spellen = **{n}** teams")

        if n_solo < 0:
            st.error("Te veel duo-spellen voor het aantal teams.")
            return

        st.divider()

        # Duo game names
        if n_duo > 0:
            st.subheader("Duo-spelnamen")
            use_def_duo = st.toggle("Standaard duo-namen", value=True, key="tog_duo")
            duo_names: list[str] = []
            if use_def_duo:
                duo_names = [f"Duo spel {i + 1}" for i in range(n_duo)]
                st.caption(", ".join(duo_names))
            else:
                for i in range(n_duo):
                    nm = st.text_input(f"Duo spel {i + 1}", value=f"Duo spel {i + 1}", key=f"duo_{i}")
                    duo_names.append(nm.strip() or f"Duo spel {i + 1}")
            st.divider()
        else:
            duo_names = []

        # Solo game names
        if n_solo > 0:
            st.subheader("Solo-spelnamen")
            use_def_solo = st.toggle("Standaard solo-namen", value=True, key="tog_solo")
            solo_names: list[str] = []
            if use_def_solo:
                solo_names = [f"Solo spel {i + 1}" for i in range(n_solo)]
                st.caption(", ".join(solo_names))
            else:
                for i in range(n_solo):
                    nm = st.text_input(f"Solo spel {i + 1}", value=f"Solo spel {i + 1}", key=f"solo_{i}")
                    solo_names.append(nm.strip() or f"Solo spel {i + 1}")
            st.divider()
        else:
            solo_names = []

        # Time
        st.subheader("Tijdslots")
        start_t = st.time_input("Starttijd", value=datetime.strptime("09:00", "%H:%M").time())
        game_min = int(st.number_input(
            "Speeltijd per ronde (min)", min_value=1, max_value=120, value=20, step=1
        ))
        switch_min = int(st.number_input(
            "Wisseltijd tussen rondes (min)", min_value=0, max_value=60, value=5, step=1
        ))
        total_min = n * (game_min + switch_min) - switch_min
        start_dt = datetime.combine(datetime.today(), start_t)
        end_dt = start_dt + timedelta(minutes=total_min)
        st.caption(
            f"Eindtijd: **{end_dt.strftime('%H:%M')}** · Totaal: **{total_min} min**"
        )

        st.divider()

        seed = int(st.number_input(
            "Variatie (seed)", min_value=0, max_value=9999, value=42, step=1,
            help="Wijzig dit getal voor een ander schema met dezelfde overige instellingen."
        ))
        gen = st.button("Genereer schema", type="primary", use_container_width=True)

    # ── Main ──────────────────────────────────────────────────────────────────
    try:
        c1, c2 = st.columns([1, 10])
        with c1:
            st.image("logo-stiphout-vooruit.svg", width=70)
        with c2:
            st.title("Voetbalkamp Schema Generator")
            st.caption("Latin Square — elk team speelt elk spel precies 1× per ronde")
    except Exception:
        st.title("Voetbalkamp Schema Generator")

    # When generate is clicked, compute and store in session_state
    if gen:
        # Validate
        errors = []
        if len(set(team_names)) < n:
            errors.append("Teamnamen moeten uniek zijn.")
        all_game_names = duo_names + solo_names
        if len(set(all_game_names)) < len(all_game_names):
            errors.append("Spelnamen moeten uniek zijn.")
        for err in errors:
            st.error(err)
        if errors:
            return

        # Build and store in session_state
        square = generate_latin_square(n, seed=seed)
        square = optimize_column_pairing(square, n_duo, n_solo, seed=seed)
        slot_labels = make_time_labels(n, start_dt, game_min, switch_min)
        df = build_dataframe(square, team_names, duo_names, solo_names, slot_labels, n_duo, n_solo)

        st.session_state["square"] = square
        st.session_state["slot_labels"] = slot_labels
        st.session_state["df"] = df
        st.session_state["n_duo"] = n_duo
        st.session_state["n_solo"] = n_solo
        st.session_state["solo_names"] = solo_names
        st.session_state["duo_names"] = duo_names
        st.session_state["team_names"] = team_names
        st.session_state["n"] = n
        st.session_state["end_dt"] = end_dt

    # If nothing generated yet, show info
    if "df" not in st.session_state:
        st.info("Vul de instellingen in de sidebar in en klik op **Genereer schema**.")
        return

    # Read from session_state
    square = st.session_state["square"]
    slot_labels = st.session_state["slot_labels"]
    df = st.session_state["df"]
    n_duo = st.session_state["n_duo"]
    n_solo = st.session_state["n_solo"]
    solo_names = st.session_state["solo_names"]
    duo_names = st.session_state["duo_names"]
    team_names = st.session_state["team_names"]
    n = st.session_state["n"]
    end_dt = st.session_state["end_dt"]

    # Metrics
    cols = st.columns(5)
    cols[0].metric("Teams", n)
    cols[1].metric("Rondes", n)
    cols[2].metric("Duo-spellen", n_duo)
    cols[3].metric("Solo-spellen", n_solo)
    cols[4].metric("Eindtijd", end_dt.strftime("%H:%M"))

    st.divider()
    st.subheader("Schema")

    # Highlight solo columns
    def highlight_solo(df):
        styles = pd.DataFrame("", index=df.index, columns=df.columns)
        for col in solo_names:
            if col in df.columns:
                styles[col] = "background-color: #fff3cd; color: #7d5800; font-weight: 500"
        return styles

    if n_solo > 0:
        st.dataframe(
            df.style.apply(highlight_solo, axis=None),
            use_container_width=True,
            height=min(40 + 36 * n, 600),
        )
    else:
        st.dataframe(df, use_container_width=True, height=min(40 + 36 * n, 600))

    st.divider()

    # Game descriptions
    st.subheader("Spelbeschrijvingen")
    all_game_names = list(duo_names) + list(solo_names)
    game_descriptions = {}
    game_tabs = st.tabs(all_game_names)
    for tab, gname in zip(game_tabs, all_game_names):
        with tab:
            desc = st.text_area(
                f"Beschrijving voor {gname}",
                value=st.session_state.get(f"desc_{gname}", ""),
                key=f"desc_input_{gname}",
                height=120,
                placeholder="Voer hier de spelregels / beschrijving in...",
                label_visibility="collapsed",
            )
            st.session_state[f"desc_{gname}"] = desc
            game_descriptions[gname] = desc

    st.divider()

    # Per team view
    with st.expander("Schema per team bekijken"):
        tabs = st.tabs(team_names)
        for tab, tm_name in zip(tabs, team_names):
            tm_idx = team_names.index(tm_name)
            with tab:
                rows = []
                for i, label in enumerate(slot_labels):
                    pos = square[i].index(tm_idx)
                    # Determine which game this position maps to
                    if pos < 2 * n_duo:
                        game_idx = pos // 2
                        partner_pos = pos + 1 if pos % 2 == 0 else pos - 1
                        partner = team_names[square[i][partner_pos]]
                        activity = f"{duo_names[game_idx]}: vs {partner}"
                    else:
                        solo_idx = pos - 2 * n_duo
                        activity = solo_names[solo_idx]
                    rows.append({"Tijdslot": label, "Activiteit": activity})
                st.dataframe(pd.DataFrame(rows), use_container_width=True, hide_index=True)

    # Verification
    with st.expander("Verificatie"):
        ok = True
        for i in range(n):
            if sorted(square[i]) != list(range(n)):
                st.error(f"Rij {i} klopt niet")
                ok = False
        for j in range(n):
            if sorted(square[i][j] for i in range(n)) != list(range(n)):
                st.error(f"Kolom {j} klopt niet")
                ok = False
        if ok:
            st.success(
                "Latin Square geverifieerd: elk team speelt in elke ronde precies 1×, "
                "en elk team komt op elke positie precies 1× voor."
            )

        # Show the Latin Square with 1-based team numbers and row/col sums
        st.subheader("Latin Square (teamnummers)")
        expected_sum = n * (n + 1) // 2
        st.caption(f"Verwachte som per rij en kolom: **1 + 2 + ... + {n} = {expected_sum}**")

        # Build display: 1-based team numbers
        ls_data = {}
        for j in range(n):
            ls_data[f"Pos {j + 1}"] = [square[i][j] + 1 for i in range(n)]
        ls_data["Σ Rij"] = [sum(square[i][j] + 1 for j in range(n)) for i in range(n)]
        ls_df = pd.DataFrame(ls_data, index=[f"Ronde {i + 1}" for i in range(n)])

        # Add column sums as last row
        col_sums = {}
        for j in range(n):
            col_sums[f"Pos {j + 1}"] = sum(square[i][j] + 1 for i in range(n))
        col_sums["Σ Rij"] = ""
        sum_row = pd.DataFrame(col_sums, index=["Σ Kolom"])
        ls_df = pd.concat([ls_df, sum_row])

        def highlight_sums(s):
            styles = pd.DataFrame("", index=s.index, columns=s.columns)
            styles["Σ Rij"] = "background-color: #1F4E79; color: white; font-weight: bold"
            styles.iloc[-1] = "background-color: #1F4E79; color: white; font-weight: bold"
            return styles

        st.dataframe(
            ls_df.style.apply(highlight_sums, axis=None),
            use_container_width=True,
        )

    # Download
    default_name = f"spellenschema_{n}_teams_{datetime.now().strftime('%Y%m%d_%H%M')}"
    file_name = st.text_input("Bestandsnaam", value=default_name, key="file_name")
    if not file_name.strip():
        file_name = default_name
    st.download_button(
        "Download als Excel",
        data=to_excel(df, n_duo, square, slot_labels, team_names, duo_names, solo_names, n_solo, game_descriptions),
        file_name=f"{file_name.strip()}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True,
    )


if __name__ == "__main__":
    main()
